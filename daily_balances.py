#!/usr/bin/env python
# coding: utf-8

# In[ ]:

import os
import openpyxl
import datetime
import win32com.client
from openpyxl.utils import column_index_from_string
# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
mapi = outlook.GetNamespace("MAPI")
inbox = mapi.GetDefaultFolder(6)  # 6 corresponds to the Inbox folder

# Search for the most recent email with the subject "Daily Cash Balances" and containing an Excel attachment
found_attachment = False
latest_attachment_date = None
latest_attachment_filename = None

messages = inbox.Items
messages.Sort("[ReceivedTime]", True)  # Sort by received time in descending order

for message in messages:
    if message.Subject == "Daily Cash Balances":
        attachments = message.Attachments
        for attachment in attachments:
            if attachment.FileName.lower().endswith(".xlsx"):
                found_attachment = True
                if latest_attachment_date is None or message.ReceivedTime.date() > latest_attachment_date:
                    latest_attachment_date = message.ReceivedTime.date()
                    latest_attachment_filename = attachment.FileName
        if found_attachment:
            break

if not found_attachment:
    print("No suitable attachment found.")
else:
    # Create the folder if it doesn't exist
    folder_path = r"U:\Python\Cash Balances\Daily_Balances"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    # Save the latest attachment to the folder
    latest_attachment_path = os.path.join(folder_path, latest_attachment_filename)
    attachment.SaveAsFile(latest_attachment_path)
    print(f"Latest attachment '{latest_attachment_filename}' downloaded successfully.")

    # Load the Excel workbook
    extracted_workbook = openpyxl.load_workbook(latest_attachment_path)
    extracted_worksheet = extracted_workbook.active  # Assume data is in the active sheet

    # Extract data from the Excel file, starting from row 2 (to skip the header row)
    extracted_data = []
    for row in extracted_worksheet.iter_rows(min_row=2, values_only=True):
        # Process each row of data
        extracted_data.append(row)

    # Define the folder and filename of the master Excel sheet
    master_folder_path = r"U:\Python\Cash Balances"
    master_filename = "February 2024 Daily Account Cash Balance.xlsx"
    master_file_path = os.path.join(master_folder_path, master_filename)

    # Get the current date and determine the dates for the previous and previous previous days, excluding weekends
    current_date = datetime.date.today()

    # Logic to handle previous and previous previous days
    if current_date.weekday() == 0:  # If today is Monday, use data from Friday
        previous_date = current_date - datetime.timedelta(days=3)  # Friday
        previous_previous_date = current_date - datetime.timedelta(days=4)  # Thursday
    else:
        previous_date = current_date - datetime.timedelta(days=1)  # Yesterday
        previous_previous_date = current_date - datetime.timedelta(days=2)  # Two days ago

    # Ensure previous previous day is a weekday
    if previous_previous_date.weekday() >= 5:  # Saturday or Sunday
        previous_previous_date -= datetime.timedelta(days=previous_previous_date.weekday() - 4)

    # Convert dates to string format
    previous_day_str = previous_date.strftime("%m.%d.%Y")
    previous_previous_day_str = previous_previous_date.strftime("%m.%d.%Y")

    # Load the master Excel workbook
    master_workbook = openpyxl.load_workbook(master_file_path)

    # Check if the worksheet exists, if not, create a new one
    if previous_day_str not in master_workbook.sheetnames:
        master_workbook.create_sheet(previous_day_str)

    # Select the worksheet to work with
    current_worksheet = master_workbook[previous_day_str]

    # Clear the contents of columns A to G in the master sheet, while preserving the rest
    for row in current_worksheet.iter_rows(min_row=2, max_row=current_worksheet.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.value = None

    # Append the extracted data to the master sheet
    for i, row_data in enumerate(extracted_data):
        for j, value in enumerate(row_data):
            current_worksheet.cell(row=i + 2, column=j + 1, value=value)

    # Define the formula to update for column J
    updated_formula_J = f"=D{{row}} - INDEX('{previous_previous_day_str}'!$D:$D, MATCH({previous_day_str}!A{{row}}, '{previous_previous_day_str}'!$A:$A, 0))"

    # Update the formula in column J starting from J2
    column_index_J = column_index_from_string('J')  # Get the column index for column J
    for row in range(2, current_worksheet.max_row + 1):
        updated_formula_row_J = updated_formula_J.format(row=row)
        current_worksheet.cell(row=row, column=column_index_J, value=updated_formula_row_J)

    # Define the formula to update for column U
    updated_formula_U = f"=T{{row}} - '{previous_previous_day_str}'!T{{row}}"

    # Update the formula in column U for specified rows
    column_index_U = column_index_from_string('U')  # Get the column index for column U
    for row in [2, 3, 4, 5, 9, 10, 11, 12, 13]:  # Update specified rows in column U
        updated_formula_row_U = updated_formula_U.format(row=row)
        current_worksheet.cell(row=row, column=column_index_U, value=updated_formula_row_U)

    # Save the master Excel sheet
    master_workbook.save(master_file_path)
