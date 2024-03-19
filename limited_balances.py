#!/usr/bin/env python
# coding: utf-8

# test

# In[6]:


import os
import openpyxl
import datetime
import win32com.client
import shutil
from openpyxl.utils import column_index_from_string

# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
mapi = outlook.GetNamespace("MAPI")
inbox = mapi.GetDefaultFolder(6)  # 6 corresponds to the Inbox

# Search for the most recent email with the subject "Daily Cash Balances" containing an Excel attachment
found_attachment = False
latest_attachment_date = None
latest_attachment_filename = None
latest_attachment_path = None  # Initialize variable to store the attachment path

messages = inbox.Items
messages.Sort("[ReceivedTime]", True)  # Sort by received time in descending order

for message in messages:
    if message.Subject == "Daily Cash Balances" and message.Attachments.Count > 0:
        attachments = message.Attachments
        for attachment in attachments:
            if attachment.FileName.lower().endswith(".xlsx"):
                found_attachment = True
                if latest_attachment_date is None or message.ReceivedTime.date() > latest_attachment_date:
                    latest_attachment_date = message.ReceivedTime.date()
                    latest_attachment_filename = attachment.FileName
                    # Prepare the folder for saving the attachment
                    folder_path = r"J:\Python\Cash Balances\Limited_Balances"
                    if not os.path.exists(folder_path):
                        os.makedirs(folder_path)
                    latest_attachment_path = os.path.join(folder_path, latest_attachment_filename)
                    attachment.SaveAsFile(latest_attachment_path)
                    print(f"Latest attachment '{latest_attachment_filename}' downloaded successfully.")
        if found_attachment:
            break

if not found_attachment or latest_attachment_path is None:
    print("No suitable attachment found.")
else:
    # Load the Excel workbook
    extracted_workbook = openpyxl.load_workbook(latest_attachment_path)
    extracted_worksheet = extracted_workbook.active  # Assume data is in the active sheet

    # Extract data from the Excel file, starting from row 2 (to skip the header row)
    filtered_data = []
    for row in extracted_worksheet.iter_rows(min_row=2, values_only=True):
        if row[5] == "L":  # Assuming "POOL" column is F (index 5)
            filtered_data.append(row)
    print(f"Filtered {len(filtered_data)} rows of data.")
    if not filtered_data:
        print("No data matching the filter criteria found. Exiting...")
    else:
        filtered_data.sort(key=lambda x: x[3], reverse=True)

    # Define the folder and filename of the master Excel sheet
    master_folder_path = r"J:\Investments\Kim\eMars Account\FY24\Limited\Daily"
    master_filename = "March 2024 Daily Limited Account Cash Balance.xlsx"
    master_file_path = os.path.join(master_folder_path, master_filename)

    # Before making changes to the master file, create a backup with a timestamp
    backup_folder_path = r"J:\Python\Cash Balances\Backups\Limited"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_filename = f"{os.path.splitext(master_filename)[0]}_backup_{timestamp}{os.path.splitext(master_filename)[1]}"
    backup_file_path = os.path.join(backup_folder_path, backup_filename)
    shutil.copy(master_file_path, backup_file_path)
    print(f"Backup created: {backup_file_path}")

    # Get the current date and determine the dates for the previous and previous previous days, excluding weekends
    current_date = datetime.date.today()

    if current_date.weekday() == 0:  # If today is Monday, use data from Friday
        previous_date = current_date - datetime.timedelta(days=3)
        previous_previous_date = current_date - datetime.timedelta(days=4)
    else:
        previous_date = current_date - datetime.timedelta(days=1)
        previous_previous_date = current_date - datetime.timedelta(days=2)

    if previous_previous_date.weekday() >= 5:  # Adjust for weekends
        previous_previous_date -= datetime.timedelta(days=previous_previous_date.weekday() - 4)

    previous_day_str = f"{previous_date.month}.{previous_date.day}.{previous_date.year}"
    previous_previous_day_str = previous_previous_date.strftime("%m.%d.%Y")

    # Load the master Excel workbook
    master_workbook = openpyxl.load_workbook(master_file_path)
    
    if previous_day_str not in master_workbook.sheetnames:
        master_workbook.create_sheet(previous_day_str)
    current_worksheet = master_workbook[previous_day_str]

    # Clear the contents of columns A to G in the master sheet
    for row in current_worksheet.iter_rows(min_row=2, max_row=current_worksheet.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.value = None

    # Append the extracted data to the master sheet
    for i, row_data in enumerate(filtered_data):
        for j, value in enumerate(row_data):
            current_worksheet.cell(row=i + 2, column=j + 1, value=value)

    # Save the master Excel sheet
    master_workbook.save(master_file_path)
    print(f"Master file '{master_filename}' has been updated and saved.")


# #j connections: 
# Backup spots: J:\Python\Cash Balances\Backups\Daily and J:\Python\Cash Balances\Backups\Limited
# Main copy spot: J:\Investments\Kim\eMars Account\FY24\Daily and J:\Investments\Kim\eMars Account\FY24\Limited\Daily
# 
# u connections for testing:
# Test documents: U:\Python\Cash Balances (code in here dunno if it works?)
